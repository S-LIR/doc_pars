from fastapi import FastAPI, File, UploadFile, Request, HTTPException, status, Depends, Security
from fastapi.concurrency import run_in_threadpool
from fastapi.responses import JSONResponse
from typing import List, Optional
import tempfile
import shutil
import os
import re
import io
from bs4 import BeautifulSoup
from email.parser import BytesParser
from email import policy
from striprtf.striprtf import rtf_to_text

from pdfminer.high_level import extract_text
from docx import Document
from openpyxl import load_workbook
import xlrd
import extract_msg
import base64
from pydantic import BaseModel
from pypdf import PdfMerger, PdfReader, PdfWriter
from dotenv import load_dotenv
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials

app = FastAPI()

MAX_TEXT_LENGTH = 10000
load_dotenv()
_API_TOKEN = os.getenv("API_TOKEN")

security = HTTPBearer(auto_error=False)


def verify_token(credentials: HTTPAuthorizationCredentials = Security(security)) -> None:
    if not _API_TOKEN:
        raise HTTPException(status_code=500, detail="API token is not configured")
    if credentials is None or not credentials.credentials:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Unauthorized")
    token = credentials.credentials.strip()
    if token != _API_TOKEN:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Forbidden")


@app.get("/test")
def test(_: None = Depends(verify_token)):
    return {"message": "сервер работает"}


def parse_pdf(path: str) -> str:
    try:
        return extract_text(path)
    except Exception as e:
        return f"[Ошибка при чтении PDF]: {str(e)}"


def parse_docx(path: str) -> str:
    try:
        doc = Document(path)
        return "\n".join(p.text for p in doc.paragraphs)
    except Exception as e:
        return f"[Ошибка при чтении DOCX]: {str(e)}"


def parse_xlsx(path: str) -> str:
    wb = None
    try:
        wb = load_workbook(path, data_only=True, read_only=True)
        sheet = wb.active
        rows = []
        for row in sheet.iter_rows(values_only=True):
            line = "\t".join(
                str(cell) if cell is not None else "" for cell in row)
            rows.append(line)
        return "\n".join(rows)
    except Exception as e:
        return f"[Ошибка при чтении XLSX]: {str(e)}"
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


def parse_txt(path: str) -> str:
    try:
        with open(path, encoding="utf-8") as f:
            return f.read()
    except Exception as e:
        return f"[Ошибка при чтении TXT]: {str(e)}"


def parse_xls(path: str) -> str:
    try:
        wb = xlrd.open_workbook(path)
        sheet = wb.sheet_by_index(0)
        rows = []
        for row_idx in range(sheet.nrows):
            row = sheet.row_values(row_idx)
            line = "\t".join(str(cell) for cell in row)
            rows.append(line)
        return "\n".join(rows)
    except Exception as e:
        return f"[Ошибка при чтении XLS]: {str(e)}"


# ===== общие утилиты =====
def html_to_text(html: str) -> str:
    soup = BeautifulSoup(html, "html.parser")
    for tag in soup(["style", "script", "head", "title", "meta"]):
        tag.decompose()
    return soup.get_text("\n", strip=True)


_QUOTE_PATTERNS = [
    re.compile(r'^\s*>'),
    re.compile(r'^\s*On .+ wrote:\s*$', re.I),
    re.compile(r'^\s*From:\s*.+$', re.I),
    re.compile(r'^\s*Sent:\s*.+$', re.I),
    re.compile(r'^\s*To:\s*.+$', re.I),
    re.compile(r'^\s*Subject:\s*.+$', re.I),
    re.compile(r'^\s*От:\s*.+$', re.I),
    re.compile(r'^\s*Кому:\s*.+$', re.I),
    re.compile(r'^\s*Тема:\s*.+$', re.I),
    re.compile(r'^\s*Отправлено:\s*.+$', re.I),
    re.compile(r'.+написал\(а\):\s*$', re.I),
]

_SIGNATURE_START = re.compile(
    r'(?im)^(?:--\s*$|—\s*$|С уважением[,!]*|С наилучшими пожеланиями[,!]*|Best regards[,]?!*|Kind regards[,]?!*|Regards[,]?!*)'
)

_DISCLAIMER_START = re.compile(
    r'(?is)(?:\bDISCLAIMER\b|This e[- ]?mail.*confidential|Это сообщение может содержать конфиденциальную информацию)'
)


def clean_email_body(text: str) -> str:
    lines = text.splitlines()
    cleaned = []
    for line in lines:
        if any(pat.match(line) for pat in _QUOTE_PATTERNS):
            break
        cleaned.append(line)
    text = "\n".join(cleaned).strip()

    m = _SIGNATURE_START.search(text)
    if m:
        text = text[:m.start()].rstrip()

    m2 = _DISCLAIMER_START.search(text)
    if m2 and len(text) - m2.start() > 200:
        text = text[:m2.start()].rstrip()

    text = re.sub(r'\n{3,}', '\n\n', text)
    return text

# ===== EML =====
def extract_eml_body(msg) -> str:
    # Предпочитаем text/plain, игнорируя явные вложения
    if msg.is_multipart():
        html_candidates = []
        for part in msg.walk():
            ctype = part.get_content_type()
            disp = part.get_content_disposition()
            if disp in ("attachment", "inline") and part.get_filename():
                continue
            if ctype == "text/plain":
                return part.get_content()
            if ctype == "text/html":
                html_candidates.append(part.get_content())
        if html_candidates:
            return html_to_text(html_candidates[0])
        return ""
    else:
        ctype = msg.get_content_type()
        content = msg.get_content()
        if ctype == "text/html":
            return html_to_text(content)
        return content or ""


def parse_eml(path: str) -> str:
    try:
        with open(path, "rb") as f:
            msg = BytesParser(policy=policy.default).parse(f)
        body = extract_eml_body(msg)
        return clean_email_body(body)
    except Exception as e:
        return f"[Ошибка при чтении EML]: {str(e)}"


# ===== MSG =====
def parse_msg(path: str) -> str:
    msg = None
    try:
        msg = extract_msg.Message(path)

        # 1) Пытаемся plain текст
        text = (msg.body or "").strip()

        # 2) Если пусто — пробуем HTML
        if not text:
            html = getattr(msg, "htmlBody", None) or getattr(
                msg, "bodyHTML", None)
            if html:
                text = html_to_text(html)

        # 3) Если всё ещё пусто — пробуем RTF
        if not text:
            rtf = getattr(msg, "rtfBody", None) or getattr(msg, "rtf", None)
            if rtf:
                try:
                    text = rtf_to_text(rtf).strip()
                except Exception:
                    # как fallback — уберём управляющие группы грубым способом
                    text = re.sub(
                        r'\\[a-z]+-?\d* ?|{\*?\\[^}]*}|[{}]', '', rtf)
                    text = re.sub(r'\s+', ' ', text).strip()

        return clean_email_body(text)
    except Exception as e:
        return f"[Ошибка при чтении MSG]: {str(e)}"
    finally:
        if msg is not None:
            try:
                msg.close()
            except Exception:
                pass


def parse_file(path: str) -> str:
    ext = os.path.splitext(path)[1].lower()
    if ext == ".pdf":
        return parse_pdf(path)
    elif ext == ".docx":
        return parse_docx(path)
    elif ext == ".xlsx":
        return parse_xlsx(path)
    elif ext == ".xls":
        return parse_xls(path)
    elif ext == ".msg":
        return parse_msg(path)
    elif ext == ".eml":
        return parse_eml(path)
    elif ext == ".txt":
        return parse_txt(path)
    else:
        return f"[Формат {ext} не поддерживается]"


def _tempfile_params(filename: Optional[str]) -> dict:
    suffix = ""
    if filename:
        _, ext = os.path.splitext(filename)
        if ext and "/" not in ext and "\\" not in ext:
            suffix = ext
    return {"prefix": "upload_", "suffix": suffix or ".tmp", "delete": False}


def _is_pdf_bytes(payload: bytes) -> bool:
    # Простая проверка по сигнатуре PDF
    return len(payload) >= 5 and payload[:5] == b"%PDF-"


def _merge_pdfs(pdf_blobs: List[bytes]) -> bytes:
    merger = PdfMerger()
    try:
        for blob in pdf_blobs:
            merger.append(io.BytesIO(blob))
        buffer = io.BytesIO()
        merger.write(buffer)
        return buffer.getvalue()
    finally:
        merger.close()


def _split_pdf_pages(pdf_blob: bytes) -> List[bytes]:
    try:
        reader = PdfReader(io.BytesIO(pdf_blob), strict=False)
    except Exception as exc:
        # Преобразуем низкоуровневую ошибку парсинга в понятную клиенту
        raise ValueError(f"Некорректный или повреждённый PDF: {exc}")

    if not reader.pages:
        return [pdf_blob]

    pages: List[bytes] = []
    for page in reader.pages:
        writer = PdfWriter()
        writer.add_page(page)
        buffer = io.BytesIO()
        writer.write(buffer)
        pages.append(buffer.getvalue())
    return pages


@app.post("/parse")
async def parse_files(files: List[UploadFile] = File(...), _: None = Depends(verify_token)):
    results = []

    for upload in files:
        params = _tempfile_params(upload.filename)
        tmp_path = None
        try:
            with tempfile.NamedTemporaryFile(**params) as tmp:
                tmp_path = tmp.name
                shutil.copyfileobj(upload.file, tmp)

            try:
                text = await run_in_threadpool(parse_file, tmp_path)
            except Exception as exc:
                text = f"[Ошибка при обработке файла]: {exc}"
            results.append({
                "filename": upload.filename,
                # можно убрать ограничение, если нужно
                "text": text[:MAX_TEXT_LENGTH]
            })
        finally:
            if tmp_path and os.path.exists(tmp_path):
                try:
                    os.unlink(tmp_path)
                except OSError:
                    pass

    return JSONResponse(content={"parsed_files": results})


class FileB64(BaseModel):
    filename: Optional[str] = None
    filedata: str


class PdfFileB64(BaseModel):
    filedata: str


class PdfEditRequest(BaseModel):
    split: bool = False
    data: List[PdfFileB64]


@app.post("/parse_base64")
def parse_file_base64(item: FileB64, _: None = Depends(verify_token)):
    try:
        data = base64.b64decode(item.filedata)
        params = _tempfile_params(item.filename)
        tmp_path = None
        try:
            with tempfile.NamedTemporaryFile(**params) as tmp:
                tmp_path = tmp.name
                tmp.write(data)

            text = parse_file(tmp_path)
        finally:
            if tmp_path and os.path.exists(tmp_path):
                try:
                    os.unlink(tmp_path)
                except OSError:
                    pass
        return JSONResponse(content={"filename": item.filename, "text": text[:MAX_TEXT_LENGTH]})
    except Exception as e:
        return JSONResponse(content={"error": str(e)}, status_code=500)


@app.post("/pdf_edit")
def pdf_edit(request: PdfEditRequest, _: None = Depends(verify_token)):
    try:
        pdf_blobs: List[bytes] = []
        for idx, item in enumerate(request.data, start=1):
            try:
                decoded = base64.b64decode(item.filedata)
            except Exception:
                raise ValueError(
                    f"Не удалось декодировать PDF (элемент #{idx})")

            if not _is_pdf_bytes(decoded):
                raise ValueError(
                    f"Переданные данные (элемент #{idx}) не являются PDF")

            pdf_blob = decoded
            pdf_blobs.append(pdf_blob)

        if not pdf_blobs:
            return JSONResponse(content={"data": []})

        if request.split:
            parts: List[bytes] = []
            for blob in pdf_blobs:
                parts.extend(_split_pdf_pages(blob))
            response_payload = [{
                "filedata": base64.b64encode(part).decode("utf-8")
            } for part in parts]
        else:
            merged = _merge_pdfs(pdf_blobs)
            response_payload = [{
                "filedata": base64.b64encode(merged).decode("utf-8")
            }]

        return JSONResponse(content={"data": response_payload})
    except ValueError as exc:
        return JSONResponse(content={"error": str(exc)}, status_code=400)
    except Exception as exc:
        return JSONResponse(content={"error": str(exc)}, status_code=500)
